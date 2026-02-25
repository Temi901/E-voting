from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from datetime import datetime
import os


def generate_results_pdf(election):
    """
    Generate a PDF report for election results
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=12,
    )
    
    # Title
    title = Paragraph(f"<b>Election Results Report</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Election Details
    details_style = styles['Normal']
    election_info = f"""
    <b>Election:</b> {election.title}<br/>
    <b>Description:</b> {election.description}<br/>
    <b>Start Date:</b> {election.start_date.strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>End Date:</b> {election.end_date.strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>Report Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>
    """
    elements.append(Paragraph(election_info, details_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Results heading
    results_heading = Paragraph("<b>Results Summary</b>", heading_style)
    elements.append(results_heading)
    elements.append(Spacer(1, 0.2*inch))
    
    # Get candidates and votes
    candidates = election.candidates.all().order_by('-votes__id')
    total_votes = sum(candidate.vote_count() for candidate in candidates)
    
    # Create results table
    table_data = [['Rank', 'Candidate', 'Party', 'Votes', 'Percentage']]
    
    for idx, candidate in enumerate(candidates, 1):
        votes = candidate.vote_count()
        percentage = (votes / total_votes * 100) if total_votes > 0 else 0
        
        rank = f"🏆 {idx}" if idx == 1 else str(idx)
        
        table_data.append([
            rank,
            candidate.name,
            candidate.party,
            str(votes),
            f"{percentage:.1f}%"
        ])
    
    # Add total row
    table_data.append(['', '', 'TOTAL', str(total_votes), '100%'])
    
    # Create table
    table = Table(table_data, colWidths=[0.8*inch, 2.5*inch, 2*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -2), 1, colors.black),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Winner announcement
    if candidates.exists():
        winner = candidates[0]
        winner_text = f"""
        <b><font size="14" color="#27ae60">🎉 WINNER: {winner.name} ({winner.party})</font></b><br/>
        <font size="12">Total Votes: {winner.vote_count()} ({(winner.vote_count()/total_votes*100):.1f}%)</font>
        """
        elements.append(Paragraph(winner_text, styles['Normal']))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = """
    <para align="center">
    <font size="8" color="gray">
    --- End of Report ---<br/>
    E-Voting System | Secure • Transparent • Democratic<br/>
    This is an official election results report.
    </font>
    </para>
    """
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_results_excel(election):
    """
    Generate an Excel spreadsheet for election results
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Election Results"
    
    # Header styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws['A1'] = "ELECTION RESULTS REPORT"
    ws['A1'].font = Font(bold=True, size=16, color="2c3e50")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Election details
    ws['A3'] = "Election:"
    ws['B3'] = election.title
    ws['A4'] = "Description:"
    ws['B4'] = election.description
    ws['A5'] = "Start Date:"
    ws['B5'] = election.start_date.strftime('%B %d, %Y at %I:%M %p')
    ws['A6'] = "End Date:"
    ws['B6'] = election.end_date.strftime('%B %d, %Y at %I:%M %p')
    ws['A7'] = "Report Generated:"
    ws['B7'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    
    # Make labels bold
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)
    
    # Results header (row 9)
    ws['A9'] = "Rank"
    ws['B9'] = "Candidate Name"
    ws['C9'] = "Party"
    ws['D9'] = "Votes"
    ws['E9'] = "Percentage"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}9'].fill = header_fill
        ws[f'{col}9'].font = header_font
        ws[f'{col}9'].alignment = Alignment(horizontal='center')
    
    # Get candidates and votes
    candidates = election.candidates.all().order_by('-votes__id')
    total_votes = sum(candidate.vote_count() for candidate in candidates)
    
    # Fill in data
    row = 10
    for idx, candidate in enumerate(candidates, 1):
        votes = candidate.vote_count()
        percentage = (votes / total_votes * 100) if total_votes > 0 else 0
        
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = candidate.name
        ws[f'C{row}'] = candidate.party
        ws[f'D{row}'] = votes
        ws[f'E{row}'] = f"{percentage:.1f}%"
        
        # Highlight winner (first row)
        if idx == 1:
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Total row
    ws[f'A{row}'] = ""
    ws[f'B{row}'] = ""
    ws[f'C{row}'] = "TOTAL"
    ws[f'D{row}'] = total_votes
    ws[f'E{row}'] = "100%"
    
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    ws[f'D{row}'].alignment = Alignment(horizontal='center')
    ws[f'E{row}'].alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer